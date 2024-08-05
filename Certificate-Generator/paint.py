from PIL import Image, ImageDraw, ImageFont
import openpyxl
import os
import qrcode

# Path to the Excel file and the font file
path = "participation.xlsx"
font_path = "arial.ttf"  # Update this path if the font is in a different location

# Load the Excel workbook
inputWorkbook = openpyxl.load_workbook(path)
inputWorksheet = inputWorkbook.active
rows = inputWorksheet.max_row

# Extract user data from the Excel sheet
user = []
for i in range(2, rows + 1):  # Starting from 2 to skip the header row
    name = inputWorksheet.cell(row=i, column=2).value
    team = inputWorksheet.cell(row=i, column=1).value
    user.append({'name': name, 'team': team})

# Ensure the certificates directory exists
certificates_dir = os.path.join(os.getcwd(), 'certificates')
os.makedirs(certificates_dir, exist_ok=True)

# Generate and save certificates for each user
for person in user:
    image = Image.open('participation.jpg')  # Use the correct path to your image
    name = person['name']
    team = person['team']

    # Define larger fonts for more prominent text
    font_regular = ImageFont.truetype(font_path, size=100)
    font_bold = ImageFont.truetype(font_path, size=150)

    draw = ImageDraw.Draw(image)

    # Text to be added to the certificate
    line1 = "This certificate is proudly presented to"
    line2 = name
    line3 = "for participating in the hackathon event held by Infosys on 27 July 2024."

    # Calculate text positions
    image_width, image_height = image.size

    # Line 1
    bbox1 = draw.textbbox((0, 0), line1, font=font_regular)
    text_width1 = bbox1[2] - bbox1[0]
    x1 = (image_width - text_width1) / 2
    y1 = image_height / 2 - 200  # Adjust as needed for spacing
    draw.text((x1, y1), line1, font=font_regular, fill="black")

    # Line 2
    bbox2 = draw.textbbox((0, 0), line2, font=font_bold)
    text_width2 = bbox2[2] - bbox2[0]
    x2 = (image_width - text_width2) / 2
    y2 = y1 + bbox1[3] + 40  # Adjust as needed for spacing
    draw.text((x2, y2), line2, font=font_bold, fill="black")

    # Line 3
    bbox3 = draw.textbbox((0, 0), line3, font=font_regular)
    text_width3 = bbox3[2] - bbox3[0]
    x3 = (image_width - text_width3) / 2
    y3 = y2 + bbox2[3] + 40  # Adjust as needed for spacing
    draw.text((x3, y3), line3, font=font_regular, fill="black")

    # Generate QR code
    qr_data = f"Name: {name}, Team: {team}"
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill='black', back_color='white')

    # Resize QR code to be larger
    qr_img = qr_img.resize((600, 600))  # Increase size as needed

    # Paste QR code onto the certificate
    qr_x = (image_width - qr_img.width) / 2
    qr_y = y3 + bbox3[3] + 80  # Move further down from the text
    image.paste(qr_img, (int(qr_x), int(qr_y)))

    # Save the certificate
    image.save(f'{certificates_dir}/{team}_{name}.jpg')
    print(f'Generated for {team}_{name}')
