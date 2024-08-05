import openpyxl
import os
import smtplib
import getpass
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Email account credentials
sender = 'ishitaandaayush@gmail.com'
app_password = 'yvir qngs qykb vvlq'
subject = 'Hack with INFI 2024 Participation Certificate'

# Path to the Excel file with participant details
path = "participation.xlsx"

# Load the Excel workbook
inputWorkbook = openpyxl.load_workbook(path)
inputWorksheet = inputWorkbook.active
rows = inputWorksheet.max_row

# Collect user details from the Excel file
user = []
for i in range(2, rows + 1):  # Starting from 2 to skip the header row
    email = inputWorksheet.cell(row=i, column=3).value
    name = inputWorksheet.cell(row=i, column=2).value
    team = inputWorksheet.cell(row=i, column=1).value
    user.append({'email': email, 'name': name, 'team': team})

# Set up the SMTP server
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(sender, app_password)

# Send emails with the certificates
for person in user:
    try:
        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = person['email']
        msg['Subject'] = subject

        team = person['team']
        name = person['name']
        email = person['email']

        # Email body
        body = f'Hi {name},\n\nThank you for participating in the Hack with Infi 2024.\nWe hope you enjoyed it and would love to see you in future events by infosys.'
        msg.attach(MIMEText(body, 'plain'))

        # Attach the certificate
        file = f'{os.getcwd()}/certificates/{team}_{name}.jpg'
        with open(file, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={team}_{name}.jpg')
            msg.attach(part)

        # Send the email
        text = msg.as_string()
        server.sendmail(sender, email, text)
        print(f'Sent mail to {email}')
    except Exception as e:
        print(f'Error sending mail to {email}: {str(e)}')

# Close the SMTP server connection
server.quit()


#